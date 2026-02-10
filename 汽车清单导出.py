import sys
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from colorama import Fore
from openpyxl import Workbook, styles, load_workbook
from openpyxl.styles import Alignment

import tkinter as tk
from pypinyin import lazy_pinyin, Style
from tkinter import filedialog
import re
import requests
import json
import os
import time
import pandas as pd

# 品牌别名映射表
BRAND_ALIASES = {
    '丰田': ['toyota', 'ft', 'fengtian', 'ft', 'fengt'],
    '比亚迪': ['byd', 'biyadi', 'build your dreams', 'bid', 'biadi'],
    '极氪': ['zeeker', 'jk', 'jike', 'zeekr', 'zker'],
    '特斯拉': ['tesla', 'tsl', 'tesila', 'tesi', 'tls'],
    '奔驰': ['benz', 'bc', 'benchi', 'mercedes', 'mb', 'benze'],
    '宝马': ['bmw', 'bm', 'baoma', 'bimmer', 'boma'],
    '奥迪': ['audi', 'ad', 'aodi', 'aodi', 'a4', 'a6'],
    '大众': ['vw', 'volkswagen', 'dazhong', 'dazhong', 'dzw'],
    '本田': ['honda', 'bt', 'bentian', 'bent', 'hond'],
    '日产': ['nissan', 'rc', 'richan', 'nisa', 'niss'],
    '现代': ['hyundai', 'xd', 'xiandai', 'hundai', 'hynd'],
    '起亚': ['kia', 'qy', 'qiya', 'ki', 'kiaa'],
    '吉利': ['geely', 'jl', 'jili', 'geeli', 'jly'],
    '长城': ['greatwall', 'gw', 'changcheng', 'cc', 'gwm'],
    '长安': ['changan', 'ca', 'chana', 'chan', 'cac'],
    '蔚来': ['nio', 'wl', 'weilai', 'nio', 'weila'],
    '小鹏': ['xpeng', 'xp', 'xiaopeng', 'xpen', 'xpng'],
    '理想': ['li', 'lixiang', 'lixiangauto', 'lix', 'lx'],
    'AC Schnitzer': ['acs', 'ac', 'schnitzer', 'ac-s', 'acsh'],
    '保时捷': ['porsche', 'bsj', 'baoshijie', 'por', 'porsh'],
    '路虎': ['landrover', 'lh', 'luhu', 'lr', 'land'],
    '捷豹': ['jaguar', 'jb', 'jiebao', 'jag', 'jagr'],
    '林肯': ['lincoln', 'lk', 'linken', 'lin', 'linc'],
    '沃尔沃': ['volvo', 'wew', 'woerwo', 'vol', 'volv'],
    '玛莎拉蒂': ['maserati', 'msld', 'mashaladi', 'mas', 'msrt'],
    '法拉利': ['ferrari', 'fll', 'falali', 'fer', 'ferr'],
    '兰博基尼': ['lamborghini', 'lbjn', 'lanbojini', 'lamb', 'lambo'],
    '宾利': ['bentley', 'bl', 'binli', 'ben', 'bentl'],
    '劳斯莱斯': ['rollsroyce', 'lsls', 'laosilaisi', 'rr', 'rolls'],
    '阿斯顿马丁': ['astonmartin', 'asdmt', 'asidunmading', 'am', 'aston'],
    '迈凯伦': ['mclaren', 'mkl', 'maikailun', 'mcl', 'mclr'],
    '阿尔法罗密欧': ['alfaromeo', 'afl', 'aerfaluomiou', 'alfa', 'alf'],
    '红旗': ['hongqi', 'hq', 'redflag', 'hongq', 'hong'],
    '领克': ['lynkco', 'lk', 'lingke', 'lynk', 'link'],
    '魏牌': ['wey', 'wp', 'weipai', 'wei', 'w牌'],
    '星途': ['exeed', 'xt', 'xingtu', 'exd', 'x途'],
    '捷途': ['jetour', 'jt', 'jietu', 'jet', 'j途'],
    '欧拉': ['ora', 'ol', 'oula', 'or', 'ola'],
    '高合': ['hiphi', 'gh', 'gaohe', 'hi', 'hiph'],
    '岚图': ['voyah', 'lt', 'lantu', 'voy', 'voya'],
    '阿维塔': ['avatr', 'awt', 'aweita', 'ava', 'avtr'],
    '智己': ['im', 'zj', 'zhiji', 'imo', 'im智'],
    '飞凡': ['rising', 'ff', 'feifan', 'ris', 'risi'],
    '哪吒': ['nezha', 'nz', 'nez', 'nzh', 'n吒'],
    '零跑': ['leapmotor', 'lp', 'lingpao', 'leap', 'l跑'],
    '威马': ['wm', 'weima', 'wmotor', 'wima', 'w马'],
    '创维': ['skyworth', 'cw', 'chuangwei', 'sky', 'skw'],
    '宝骏': ['baojun', 'bj', 'bjun', 'bao', 'b骏'],
    '五菱': ['wuling', 'wl', 'wlin', 'wu', 'w菱'],
    '名爵': ['mg', 'mj', 'mingjue', 'm爵', 'mg名'],
    '荣威': ['roewe', 'rw', 'rongwei', 'roe', 'r威'],
    '传祺': ['gac', 'cq', 'chuanqi', 'ga', 'g祺'],
    '启辰': ['venucia', 'qc', 'qichen', 'ven', 'v启'],
    '东风': ['df', 'dongfeng', 'dfeng', 'dong', 'd风'],
    '北汽': ['baic', 'bq', 'beigi', 'b汽', 'bei']
}


# 获取汽车品牌响应数据的函数
def get_band_response(brand_id="0", max_retries=3, timeout=10):
    # 在打包环境下直接使用固定用户代理
    if getattr(sys, 'frozen', False):
        headers = {"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
    else:
        try:
            # 设置缓存路径避免 fake_useragent 错误
            ua = UserAgent(cache=True, path=os.path.join(os.path.expanduser("~"), ".fake_useragent"))
            headers = {"user-agent": ua.random}
        except:
            # 如果 fake_useragent 失败，使用默认用户代理
            headers = {"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
    url = "https://car.autohome.com.cn/AsLeftMenu/As_LeftListNew.ashx"
    params = {"typeId": "1", "brandId": brand_id, "fctId": "0", "seriesId": "0"}
    
    for attempt in range(max_retries):
        try:
            response = requests.get(url, headers=headers, params=params, timeout=timeout)
            response.raise_for_status()
            return response
        except requests.RequestException as e:
            if attempt == max_retries - 1:
                print(Fore.RED + f"请求失败: {str(e)}")
                return None
            time.sleep(1)
    return None


# 预编译正则表达式
RE_DOC_WRITELN = re.compile(r'document.writeln\("(.*)"\)')
RE_SERIES_ID = re.compile(r'/price/series-(\d+).html')

# 解析车型相关信息的函数
def parse_series(band, response):
    if not response:
        return {}
        
    html_match = RE_DOC_WRITELN.search(response.text)
    if not html_match:
        return {}
        
    soup = BeautifulSoup(html_match.group(1), "html.parser")
    series_links = soup.select(".current > dl > dd > a")
    
    series_dict = {}
    for idx, link in enumerate(series_links, 1):
        series_name = link.contents[0].text.strip()
        series_id = RE_SERIES_ID.search(link.get("href", ""))
        if series_id:
            series_dict[series_id.group(1)] = series_name
            print(f"{idx}. {series_name} (ID: {series_id.group(1)})")
    
    return series_dict


# 预编译正则表达式
RE_JSON_ERROR = re.compile(r'抱歉.*暂无相关数据')

# 获取单个车型配置信息响应的函数
def get_response(series_id="0", max_retries=3, timeout=10):
    session = requests.Session()
    # 在打包环境下直接使用固定用户代理
    if getattr(sys, 'frozen', False):
        session.headers.update({"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"})
    else:
        try:
            session.headers.update({"user-agent": UserAgent().random})
        except:
            session.headers.update({"user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"})
    
    url = "https://car-web-api.autohome.com.cn/car/param/getParamConf"
    params = {"mode": "1", "site": "1", "seriesid": series_id}
    
    for attempt in range(max_retries):
        try:
            response = session.get(url, params=params, timeout=timeout)
            response.raise_for_status()
            
            if RE_JSON_ERROR.search(response.text):
                print(Fore.YELLOW + "该系列车暂无配置信息")
                return None
                
            return response
        except requests.RequestException as e:
            if attempt == max_retries - 1:
                print(Fore.RED + f"请求失败: {str(e)}")
                return None
            time.sleep(1)
    return None


def main():
    # 处理PyInstaller打包后的输入问题
    if getattr(sys, 'frozen', False):
        from tkinter import simpledialog
        root = tk.Tk()
        root.withdraw()

    while True:  # 外层循环，用于控制是否继续下载其他品牌配置
        while True:
            if getattr(sys, 'frozen', False):
                band = simpledialog.askstring("输入", "请输入汽车品牌：")
                if band is None:
                    band = ""
            else:
                band = input("请输入汽车品牌：").strip()
            response = get_band_response()
            # 获取所有品牌信息
            all_brands = re.findall(r'<a href=([^>]*?)><i[^>]*?></i>([^<]+)<em>', response.text)
            
            # 模糊匹配：检查输入是否匹配品牌名、拼音、别名或缩写
            matched_band = None
            for href, name in all_brands:
                # 检查精确匹配
                if band.lower() == name.lower():
                    matched_band = (href, name)
                    break
                
                # 检查别名映射
                aliases = BRAND_ALIASES.get(name, [])
                if band.lower() in [a.lower() for a in aliases]:
                    matched_band = (href, name)
                    break
                
                # 检查拼音全拼
                pinyin = ''.join(lazy_pinyin(name, style=Style.NORMAL))
                if band.lower() in pinyin:
                    matched_band = (href, name)
                    break
                
                # 检查拼音首字母
                initials = ''.join([x[0] for x in lazy_pinyin(name, style=Style.FIRST_LETTER)])
                if band.lower() == initials.lower():
                    matched_band = (href, name)
                    break
                
                # 检查部分匹配
                if band.lower() in name.lower():
                    matched_band = (href, name)
                    break
            
            if not matched_band:
                print("未找到匹配品牌，请重新输入")
                continue
            else:
                band_href, actual_name = matched_band
                print(f"匹配到品牌: {actual_name}")
                band_id = re.findall(r'/price/brand-(\d+).html', band_href)[0]
                print(F"{band} 品牌id为：", band_id)
                resp_brand = get_band_response(brand_id=band_id)
                # 解析车型信息，获取series_dict
                series_dict = parse_series(band, resp_brand)

                # 让用户选择保存路径
                root = tk.Tk()
                root.withdraw()  # 隐藏主窗口
                save_path = filedialog.askdirectory(
                    title="选择保存路径",
                    initialdir=os.path.join(os.path.expanduser("~"), "Downloads")
                )
                
                if not save_path:  # 用户取消选择
                    save_path = os.path.join(os.path.expanduser("~"), "Downloads", "汽车配置")
                    print(Fore.YELLOW + f"将使用默认路径: {save_path}")
                
                # 自动创建品牌文件夹
                os.makedirs(save_path, exist_ok=True)
                brand_save_path = os.path.join(save_path, actual_name)
                os.makedirs(brand_save_path, exist_ok=True)
                
                print(Fore.GREEN + f"Excel文件将保存到: {brand_save_path}")

                while True:
                    if getattr(sys, 'frozen', False):
                        choice = simpledialog.askstring("输入", "请输入需要下载的车型id，输入0则下载该品牌全部车型配置：")
                        if choice is None:
                            choice = ""
                    else:
                        choice = input(Fore.RED + "\n请输入需要下载的车型id，输入0则下载该品牌全部车型配置：").strip()
                    if choice == "0":
                        for series_id in series_dict.keys():
                            series_name = series_dict[series_id]
                            series_url = f"https://car.autohome.com.cn/config/series/{series_id}.html"
                            print(Fore.CYAN + f"正在下载 {series_name} (ID: {series_id})...")
                            
                            response = get_response(series_id)
                            if not response:
                                continue
                                
                            try:
                                resp_dict = json.loads(response.text)
                                all_info = get_car_config(resp_dict)
                                excel_name = f"{band}_{series_name}.xlsx"
                                save_to_excel(all_info, actual_name, excel_name, brand_save_path)
                            except Exception as e:
                                print(Fore.RED + f"处理 {series_name} 时出错: {str(e)}")
                        break
                    elif choice in series_dict.keys():
                        series_name = series_dict[choice]
                        series_url = f"https://car.autohome.com.cn/config/series/{choice}.html"
                        print(Fore.CYAN + f"正在下载 {series_name} (ID: {choice})...")
                        
                        response = get_response(choice)
                        if not response:
                            continue
                            
                        try:
                            resp_dict = json.loads(response.text)
                            all_info = get_car_config(resp_dict)
                            excel_name = f"{band}_{series_name}.xlsx"
                            save_to_excel(all_info, actual_name, excel_name, brand_save_path)
                        except Exception as e:
                            print(Fore.RED + f"处理 {series_name} 时出错: {str(e)}")
                        break
                    else:
                        print("输入的车型id不存在，请重新输入。")
            break

        # 询问用户是否继续下载其他品牌配置
        if getattr(sys, 'frozen', False):
            continue_choice = simpledialog.askstring("输入", "是否继续下载其他汽车品牌配置信息（Y/N）：")
            if continue_choice is None:
                continue_choice = ""
            continue_choice = continue_choice.strip().upper()
        else:
            continue_choice = input(Fore.BLUE + "是否继续下载其他汽车品牌配置信息（Y/N）：").strip().upper()
        if continue_choice != "Y":
            break
    input("请按任意键关闭程序...")


# 清洗数据
def get_car_config(config_dic):
    """
    清洗配置数据，转换为竖向表格格式（纵向导出）
    第一列是配置项名称，后续列是各车型配置值
    """
    if not config_dic or 'result' not in config_dic:
        return []
    
    titlelist = config_dic['result'].get('titlelist', [])
    datalist = config_dic['result'].get('datalist', [])
    
    # 获取配置项列表（过滤掉选配配置）
    # 按照修复后的逻辑：遍历titlelist，收集所有非选配的配置项
    config_name_list = []
    global_index_to_config_index = {}  # 全局索引 -> config_name_list索引
    
    global_index = 0
    config_index = 0
    
    for title in titlelist:
        for item in title.get('items', []):
            itemname = item.get('itemname', '')
            if itemname and '选配' not in itemname:
                # 清理配置项名称，移除特殊字符，但保留中文字符
                clean_name = re.sub(r'[^a-zA-Z0-9\u4e00-\u9fa5]', '', itemname)
                if clean_name:
                    # 不添加C_前缀，直接使用清理后的名称
                    config_name_list.append(clean_name)
                    global_index_to_config_index[global_index] = config_index
                    config_index += 1
            global_index += 1
    
    if len(config_name_list) == 0 or len(datalist) == 0:
        return []
    
    # 竖向导出：第一列是配置项名称，后续列是各车型
    result = []
    
    # 第一行：配置项名称 + 车型名称
    first_row = ['配置项']
    for data in datalist:
        # 尝试获取车型名称
        model_name = data.get('specname') or data.get('name') or ''
        first_row.append(model_name)
    result.append(first_row)
    
    # 为每个配置项创建一行
    for config_idx in range(len(config_name_list)):
        row = [config_name_list[config_idx]]
        
        # 找到该配置项对应的全局索引
        target_global_index = -1
        for global_idx, config_idx_from_map in global_index_to_config_index.items():
            if config_idx_from_map == config_idx:
                target_global_index = global_idx
                break
        
        # 遍历每个车型，获取该配置项的值
        for data in datalist:
            config_value = ''
            
            # paramconflist的索引和titlelist中所有items的全局索引是一一对应的
            paramconflist = data.get('paramconflist', [])
            
            # 如果找到了对应的全局索引，直接从paramconflist中获取
            if target_global_index >= 0 and target_global_index < len(paramconflist):
                value_item = paramconflist[target_global_index]
                
                # 跳过选配配置项（检查○符号）
                itemname = value_item.get('itemname', '')
                if itemname and '○' in itemname:
                    # 跳过选配，但保留空值
                    config_value = ''
                elif itemname and itemname != '':
                    config_value = itemname
                elif value_item.get('sublist'):
                    # 处理多值配置（如：发动机、变速箱等）
                    sub_values = []
                    for multivalue in value_item['sublist']:
                        sub_values.append(str(multivalue.get('value', '')) + str(multivalue.get('name', '')))
                    config_value = '\n'.join(sub_values)
                else:
                    # 如果既没有itemname也没有sublist，可能是空值
                    config_value = ''
            
            row.append(config_value)
        
        result.append(row)
    
    return result


def save_to_excel(data, brand_name, filename, save_path):
    """保存数据到Excel文件（竖向导出格式）
    
    Args:
        data: 要保存的数据列表，第一行是标题行（配置项 + 车型名称），后续行是配置项和对应的值
        brand_name: 品牌名称(用于显示)
        filename: 要保存的文件名
        save_path: 完整的保存路径(已包含品牌目录)
    """
    if not data or len(data) < 2:
        print(Fore.RED + "无有效配置数据可保存")
        return
    
    try:
        # 竖向导出格式：第一列是配置项名称，后续列是各车型
        # data[0] 是标题行：['配置项', '车型1', '车型2', ...]
        # data[1:] 是数据行：每行第一列是配置项名称，后续列是对应的配置值
        
        # 创建DataFrame，第一列是配置项名称，后续列是各车型
        # 第一行是标题，从第二行开始是数据
        headers = data[0] if data else []
        rows = data[1:] if len(data) > 1 else []
        
        # 创建DataFrame
        df = pd.DataFrame(rows, columns=headers)
        excel_path = os.path.join(save_path, filename)
        df.to_excel(excel_path, index=False)
        
        # 设置Excel格式
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        
        # 设置所有单元格自动换行和垂直居中
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        # 设置列宽
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 20
        
        workbook.save(excel_path)
        
        print(Fore.GREEN + f"{brand_name}配置保存完成: {excel_path}")
    except Exception as e:
        print(Fore.RED + f"保存{brand_name}配置时出错: {str(e)}")


if __name__ == '__main__':
    main()
